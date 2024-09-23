import requests
from bs4 import BeautifulSoup
import openpyxl

# Load the Excel file
file_path = 'Updated_ReplicationWiki_with_URLs.xlsx'

wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Define a function to scrape "Replication type" from the given URL
def scrape_replication_type(url):
    try:
        # Request the page content
        response = requests.get(url)
        response.raise_for_status()
        
        # Parse the HTML using BeautifulSoup
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Find the table under "This article is a replication of" and get the "Replication type" field
        table = soup.find('table', {'summary': 'This article is a replication of'})
        if table:
            replication_type = table.find('td', text='Replication type').find_next('td').get_text(strip=True)
            print(replication_type)
            return replication_type
        else:
            return "Table not found"
    
    except Exception as e:
        return f"Error: {str(e)}"

# Iterate over the rows in the worksheet
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):  # column B contains the links
    url = row[0].value  # Get the URL
    if url:
        replication_type = scrape_replication_type(url)  # Scrape the replication type
        ws.cell(row=row[0].row, column=1).value = replication_type  # Fill column A with the scraped data

# Save the updated workbook
output_path = 'ReplicationWiki_scraped3.xlsx'
wb.save(output_path)

output_path  # Return the path to the saved file
